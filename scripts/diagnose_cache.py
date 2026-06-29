from __future__ import annotations

import sqlite3
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from mbs_insite_probe import _parse_dropdown_items, get_terms, make_session

DB_PATH = ROOT / "data" / "bookstore_cache.db"
EXCEL_PATH = Path(r"C:\Users\CMInter1\Downloads\26 Summer.xlsx")
EXCEL_SHEET = "Summer 2026 Do Not Edit"


def norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().upper()


def clean_isbn(value: Any) -> str:
    text = norm(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def dump_terms() -> None:
    print("=== MBS TERMS ===")
    client, token = make_session()
    raw_terms = get_terms(client, token, verbose=False)
    terms = _parse_dropdown_items(raw_terms, "ter")
    if not terms:
        print("No terms returned.")
        return
    for term in terms:
        print(f"{term['label']} | data-id={term['raw_id']} | term_id={term['id']}")


def excel_headers(ws) -> dict[str, int]:
    headers = [cell.value for cell in ws[2]]
    return {header: index for index, header in enumerate(headers) if header}


def is_real_excel_material(row: tuple[Any, ...], idx: dict[str, int]) -> bool:
    values = " | ".join(
        norm(row[idx[column]])
        for column in ("ISBN", "TITLE", "AUTHOR", "BOOK TYPE", "REQUIRED OPTIONAL PICK ONE")
    )
    if not values.strip(" |"):
        return False
    if "NO ADOPTION" in values or "NO TEXT REQUIRED" in values:
        return False
    return True


def load_excel_materials() -> tuple[set[tuple[str, str, str]], dict[tuple[str, str, str], list[dict[str, str]]], Counter]:
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[EXCEL_SHEET]
    idx = excel_headers(ws)
    courses: set[tuple[str, str, str]] = set()
    materials_by_course: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    categories: Counter = Counter()

    for row in ws.iter_rows(min_row=3, values_only=True):
        dept = norm(row[idx["DEPT"]])
        course = norm(row[idx["COURSE"]])
        section = norm(row[idx["SECTION"]])
        if not dept or not course or not section:
            continue
        key = (dept, course, section)
        courses.add(key)
        title = norm(row[idx["TITLE"]])
        if "NO ADOPTION" in title:
            categories["No Adoption Received"] += 1
            continue
        if "NO TEXT REQUIRED" in title:
            categories["No Text Required"] += 1
            continue
        if not is_real_excel_material(row, idx):
            categories["Other/blank material"] += 1
            continue
        categories["Material rows"] += 1
        materials_by_course[key].append(
            {
                "title": norm(row[idx["TITLE"]]),
                "isbn": clean_isbn(row[idx["ISBN"]]),
                "author": norm(row[idx["AUTHOR"]]),
                "edition": norm(row[idx["EDITION"]]),
                "requirement": norm(row[idx["REQUIRED OPTIONAL PICK ONE"]]),
                "book_type": norm(row[idx["BOOK TYPE"]]),
            }
        )

    return courses, materials_by_course, categories


def load_db_materials() -> tuple[set[tuple[str, str, str]], dict[tuple[str, str, str], list[dict[str, str]]]]:
    courses: set[tuple[str, str, str]] = set()
    materials_by_course: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        for row in conn.execute("SELECT department, course_number, section FROM courses"):
            courses.add((norm(row["department"]), norm(row["course_number"]), norm(row["section"])))
        for row in conn.execute(
            """
            SELECT
                c.department, c.course_number, c.section,
                b.title, b.isbn, b.author, b.edition, b.requirement, b.format
            FROM books b
            JOIN courses c ON b.course_id = c.id
            ORDER BY c.department, c.course_number, c.section, b.title
            """
        ):
            key = (norm(row["department"]), norm(row["course_number"]), norm(row["section"]))
            materials_by_course[key].append(
                {
                    "title": norm(row["title"]),
                    "isbn": clean_isbn(row["isbn"]),
                    "author": norm(row["author"]),
                    "edition": norm(row["edition"]),
                    "requirement": norm(row["requirement"]),
                    "format": norm(row["format"]),
                }
            )
    return courses, materials_by_course


def print_materials(label: str, materials: list[dict[str, str]]) -> None:
    print(f"  {label} ({len(materials)}):")
    if not materials:
        print("    - <none>")
        return
    for item in materials:
        details = []
        for field in ("isbn", "author", "edition", "requirement", "book_type", "format"):
            if item.get(field):
                details.append(f"{field}={item[field]}")
        suffix = f" [{' | '.join(details)}]" if details else ""
        print(f"    - {item.get('title', '')}{suffix}")


def compare_cache_to_excel() -> None:
    print("\n=== CACHE VS EXCEL MATERIAL COUNT MISMATCHES ===")
    excel_courses, excel_materials, categories = load_excel_materials()
    db_courses, db_materials = load_db_materials()

    mismatches: list[tuple[tuple[str, str, str], int, int]] = []
    for key in sorted(excel_courses & db_courses):
        excel_count = len(excel_materials.get(key, []))
        db_count = len(db_materials.get(key, []))
        if excel_count != db_count:
            mismatches.append((key, excel_count, db_count))

    print(f"Excel courses: {len(excel_courses)}")
    print(f"Excel material rows: {sum(len(items) for items in excel_materials.values())}")
    print(f"Excel row categories: {dict(categories)}")
    print(f"DB courses: {len(db_courses)}")
    print(f"DB material rows: {sum(len(items) for items in db_materials.values())}")
    print(f"Courses present in both with material-count mismatch: {len(mismatches)}")

    for key, excel_count, db_count in mismatches:
        dept, course, section = key
        print("\n" + "-" * 72)
        print(f"{dept} {course} section {section}: Excel={excel_count}, DB={db_count}")
        print_materials("Excel", excel_materials.get(key, []))
        print_materials("DB", db_materials.get(key, []))


def main() -> None:
    dump_terms()
    compare_cache_to_excel()


if __name__ == "__main__":
    main()
